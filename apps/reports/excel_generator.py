import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

def render_to_excel(filename, sheet_title, data_headers, data_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')

    ws.append(data_headers)
    for col_num, header in enumerate(data_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # Row data
    for row in data_rows:
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
